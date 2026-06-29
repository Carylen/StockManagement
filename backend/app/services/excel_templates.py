"""
Pure-function XLSX builders for downloadable templates.

Each function is synchronous and CPU-bound; call via asyncio.to_thread in async routes.

Usage:
    from app.services.excel_templates import build_readiness, build_master, build_employees
    data = await asyncio.to_thread(build_master)
"""
import io
from datetime import date

import openpyxl
import openpyxl.utils
from openpyxl.styles import Alignment, Font, PatternFill
from typing import Optional


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_BG = "1F6F4C"


def _styled_header(ws, headers: list[str], bg_hex: str = _HEADER_BG, font_color: str = "FFFFFF") -> None:
    fill  = PatternFill("solid", fgColor=bg_hex)
    font  = Font(bold=True, color=font_color)
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font  = font
        cell.fill  = fill
        cell.alignment = align


def _set_col_widths(ws, widths: list[int]) -> None:
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w


# ── Readiness (daily stock upload) ───────────────────────────────────────────

def build_readiness(site: Optional[str] = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Readiness"

    headers = ["part_number", "description", "mnemonic", "commodity", "min", "max", "status", "rtt", "tbd", "total", "estimasi"]
    _styled_header(ws, headers)

    samples = [
        ["600-311-3750", "Filter Oli Engine Komatsu",  "KOM-ENG", "ENGINE",        2.0, 5.0, "AMAN",    4, 0, 4, ""],
        ["1873018",      "Air Filter Scania P460",      "SCA-BDY", "BODY",          1.0, 3.0, "WARNING", 0, 1, 1, "15/06/2026"],
        ["207-70-73181", "Seal Kit Undercarriage",      "KOM-UDR", "UNDERCARRIAGE", 1.0, 2.0, "MAX",     2, 0, 2, ""],
    ]
    for row in samples:
        ws.append(row)

    _set_col_widths(ws, [20, 40, 12, 16, 6, 6, 10, 6, 6, 8, 12])
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Info")
    notes = [
        ["Kolom",       "Keterangan"],
        ["part_number", "Wajib. Nomor part (PN). Case-insensitive."],
        ["description", "Opsional. Nama/deskripsi part."],
        ["mnemonic",    "Opsional. Kode prefix part (contoh: KOM-ENG, SCA-BDY)."],
        ["commodity",   "Opsional. Kategori komoditi (ENGINE, BODY, UNDERCARRIAGE, dll.)."],
        ["min",         "Qty minimum stok (angka desimal diperbolehkan)."],
        ["max",         "Qty maksimum stok. Harus >= min."],
        ["status",      "WARNING | AMAN | OVER | MAX — di-recompute ulang oleh backend."],
        ["rtt",         "Qty stok RTT (integer)."],
        ["tbd",         "Qty stok TBD (integer)."],
        ["total",       "Harus sama dengan rtt + tbd. Isi 0 jika tidak tahu."],
        ["estimasi",    "Opsional. Tanggal estimasi kedatangan (format: DD/MM/YYYY atau YYYY-MM-DD). Kosongkan jika tidak ada."],
        ["",            ""],
        ["Catatan",     f"Site: {site or '(sesuai akun admin yang upload)'}"],
    ]
    for r in notes:
        info.append(r)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Master Class V/G ──────────────────────────────────────────────────────────

def build_master() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Class VG"

    headers = ["Stockcode", "Part Number", "Description", "Mnemonic", "Commodity", "Class"]
    _styled_header(ws, headers)

    samples = [
        ["KOM-ENG-001", "600-311-3750",  "Filter Oli Engine Komatsu",  "KOM-ENG", "ENGINE",        "V"],
        ["SCA-BDY-001", "1873018",       "Air Filter Scania P460",     "SCA-BDY", "BODY",          "G"],
        ["KOM-UDR-001", "207-70-73181",  "Seal Kit Undercarriage",     "KOM-UDR", "UNDERCARRIAGE", "V"],
    ]
    for row in samples:
        ws.append(row)

    _set_col_widths(ws, [18, 20, 45, 16, 20, 8])
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Info")
    notes = [
        ["Kolom",       "Keterangan"],
        ["Stockcode",   "Opsional. Kode stok internal."],
        ["Part Number", "Wajib. Nomor part (PN)."],
        ["Description", "Opsional. Nama/deskripsi part."],
        ["Mnemonic",    "Opsional. Prefix KOM* → Komatsu, lainnya → Scania."],
        ["Commodity",   "Opsional. Nama komoditi (ENGINE, BODY, UNDERCARRIAGE, dll.)."],
        ["Class",       "Wajib. V atau G (huruf kapital)."],
    ]
    for r in notes:
        info.append(r)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Scheduled Plan (empty template, DELTA3 D.3) ───────────────────────────────

def build_plan_template(role: str) -> bytes:
    """Empty starter template for a scheduled-plan upload. `role="planner"`
    matches the baseline-upload columns (plan_parser.REQUIRED_COLUMNS) so a
    filled-in copy round-trips through the real parser; `role="supplier"`
    matches the fill-upload natural-key columns (no STATUS column, per
    DELTA2)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if role == "supplier":
        ws.title = "Fill Template"
        headers = ["EGI", "CN", "APL ACTIVITY", "NPN", "DESC", "REQ QTY", "REQ DATE", "UT LOCATION", "EST DATE"]
        _styled_header(ws, headers, bg_hex="E8A323")
        ws.append(["EGI-001", "CN-001", "OVERHAUL ENGINE", "600-311-3750", "Filter Oli Engine", 2, "01/07/2026", "ready", "05/07/2026"])
        _set_col_widths(ws, [14, 14, 22, 18, 30, 10, 12, 14, 12])
    else:
        ws.title = "Baseline Template"
        headers = ["DISTRIK", "EGI", "CN", "ACTIVITY", "APL ACTIVITY", "NPN", "DESC", "REQ QTY", "REQ DATE"]
        _styled_header(ws, headers)
        ws.append(["AGMR", "EGI-001", "CN-001", "OVERHAUL", "OVERHAUL ENGINE", "600-311-3750", "Filter Oli Engine", 2, "01/07/2026"])
        _set_col_widths(ws, [10, 14, 14, 12, 22, 18, 30, 10, 12])

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Employees (bulk import) ───────────────────────────────────────────────────

def build_employees() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Karyawan"

    headers = ["NRP", "Nama", "Posisi", "Shift"]
    _styled_header(ws, headers)

    samples = [
        ["KM19142", "Budi Santoso",    "User",    "Day"],
        ["KM20015", "Ahmad Fauzi",     "GL",      "Day"],
        ["GL20002", "Slamet Riyadi",   "Planner", "Day"],
        ["KM18033", "Rini Widiastuti", "User",    "Night"],
    ]
    for row in samples:
        ws.append(row)

    _set_col_widths(ws, [14, 30, 12, 10])
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Info")
    notes = [
        ["Kolom",  "Keterangan"],
        ["NRP",    "Wajib. ID karyawan (case-insensitive, disimpan uppercase)."],
        ["Nama",   "Wajib. Nama lengkap karyawan."],
        ["Posisi", "Wajib. User, GL, atau Planner (GL-Planner: bisa approve inquiry & kelola scheduled plan)."],
        ["Shift",  "Opsional. Day, Night, dsb."],
    ]
    for r in notes:
        info.append(r)
    info.column_dimensions["A"].width = 10
    info.column_dimensions["B"].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Scheduled Plan exports (data, not template) ───────────────────────────────

_EXPORT_BG = "E8A323"


def build_planner_export(rows: list, site: str) -> io.BytesIO:
    """Export all planner lines for a period. Sync — call via asyncio.to_thread."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Scheduled Plan"
    _styled_header(ws, ["DISTRIK", "EGI", "CN", "ACTIVITY", "APL ACTIVITY", "NPN", "DESC", "REQ QTY", "REQ DATE"],
                   bg_hex=_EXPORT_BG, font_color="000000")
    for i, ln in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=site)
        ws.cell(row=i, column=2, value=ln.egi)
        ws.cell(row=i, column=3, value=ln.cn)
        ws.cell(row=i, column=4, value=ln.activity)
        ws.cell(row=i, column=5, value=ln.apl_activity)
        ws.cell(row=i, column=6, value=ln.npn)
        ws.cell(row=i, column=7, value=ln.description or "")
        ws.cell(row=i, column=8, value=float(ln.req_qty) if ln.req_qty is not None else 0)
        ws.cell(row=i, column=9, value=ln.req_date.strftime("%d/%m/%Y") if ln.req_date else "")
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 45)  # type: ignore[union-attr]
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_fill_export(
    rows: list,
    site: str,
    name: str,
    due_date: date,
    days_remaining: int,
) -> io.BytesIO:
    """Export fill lines for a supplier. Sync — call via asyncio.to_thread."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fill"
    _styled_header(ws, ["EGI", "CN", "APL ACTIVITY", "NPN", "DESC", "REQ QTY", "REQ DATE", "UT LOCATION", "EST DATE"],
                   bg_hex=_EXPORT_BG, font_color="000000")
    for i, ln in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=ln.egi)
        ws.cell(row=i, column=2, value=ln.cn)
        ws.cell(row=i, column=3, value=ln.apl_activity)
        ws.cell(row=i, column=4, value=ln.npn)
        ws.cell(row=i, column=5, value=ln.description or "")
        ws.cell(row=i, column=6, value=float(ln.req_qty) if ln.req_qty is not None else 0)
        ws.cell(row=i, column=7, value=ln.req_date.strftime("%d/%m/%Y") if ln.req_date else "")
        ws.cell(row=i, column=8, value=ln.ut_location or "")
        ws.cell(row=i, column=9, value=ln.est_date.strftime("%d/%m/%Y") if ln.est_date else "")
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 45)  # type: ignore[union-attr]
    ws.freeze_panes = "A2"
    info = wb.create_sheet("Info")
    info.append(["Site", site])
    info.append(["Event", name])
    info.append(["Due date", due_date.strftime("%d/%m/%Y")])
    info.append(["Sisa hari sebelum LOCKED", days_remaining])
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Inquiry & Stock exports ───────────────────────────────────────────────────

def build_inquiry_export(rows: list[dict]) -> io.BytesIO:
    """Export inquiry lines. rows = list of flat dicts (pre-materialized in async context).
    Sync — call via asyncio.to_thread."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Inquiry Kelas G"
    _styled_header(ws, ["Tanggal", "Pemohon", "NRP", "Site",
                         "Part Number", "Part Name", "Qty", "Status", "UT Notes", "Replacement PN"],
                   bg_hex="F5A623", font_color="000000")
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["created_at"])
        ws.cell(row=i, column=2, value=r["submitter_name"])
        ws.cell(row=i, column=3, value=r["submitter_nrp"])
        ws.cell(row=i, column=4, value=r["site"])
        ws.cell(row=i, column=5, value=r["part_number"])
        ws.cell(row=i, column=6, value=r["part_name"])
        ws.cell(row=i, column=7, value=r["qty"])
        ws.cell(row=i, column=8, value=r["status"])
        ws.cell(row=i, column=9, value=r["ut_note"])
        ws.cell(row=i, column=10, value=r["replacement_pn"])
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 50)  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_stock_export(rows: list[dict], site: str) -> io.BytesIO:
    """Export stock report. rows = list of flat dicts (pre-materialized in async context).
    Sync — call via asyncio.to_thread."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"Stok {site}"
    _styled_header(ws, ["Part Number", "Deskripsi", "Komoditi", "RTT", "TBD", "Total",
                         "MIN", "MAX", "Status", "Estimasi"],
                   bg_hex="F5A623", font_color="000000")
    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r["part_number"])
        ws.cell(row=i, column=2, value=r["description"])
        ws.cell(row=i, column=3, value=r["commodity"])
        ws.cell(row=i, column=4, value=r["rtt_qty"])
        ws.cell(row=i, column=5, value=r["tbd_qty"])
        ws.cell(row=i, column=6, value=r["total_qty"])
        ws.cell(row=i, column=7, value=r["min_qty"])
        ws.cell(row=i, column=8, value=r["max_qty"])
        ws.cell(row=i, column=9, value=r["status"])
        ws.cell(row=i, column=10, value=r["estimated_date"])
    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 40)  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
