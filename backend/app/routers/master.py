"""
Master Class V/G router.

Endpoints:
  GET  /master/parts/meta      → latest upload metadata (for the Active Master card)
  GET  /master/parts           → paginated part list (all classes, no site scope)
  POST /master/parts/upload    → upload new master XLSX (admin only)
  GET  /master/parts/template  → download blank XLSX template
"""
import io
import math
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_

from app.core.database import get_db
from app.core.auth import require_user_role, Principal
from app.models.part import Part
from app.models.master_upload import MasterUpload
from app.models.user import User

router = APIRouter(prefix="/master", tags=["master"])

# ── column aliases for the master XLSX ──────────────────────────────────────
_MASTER_ALIASES: dict[str, list[str]] = {
    "no":          ["no", "#", "nomor"],
    "stockcode":   ["stockcode", "stock code", "stock_code", "kode stok"],
    "part_number": ["part number", "part_number", "part no", "pn", "new pn", "partnumber", "part no."],
    "description": ["description", "desc", "deskripsi", "nama part", "part name"],
    "mnemonic":    ["mnemonic", "mnemo", "mnemonic code"],
    "kelas":       ["class", "kelas", "kls", "classification"],
}
_REQUIRED_MASTER = {"part_number", "kelas"}


def _normalize_master_cols(df: pd.DataFrame) -> pd.DataFrame:
    col_map: dict[str, str] = {}
    upper_cols = {c.upper().strip(): c for c in df.columns}
    for canonical, aliases in _MASTER_ALIASES.items():
        for alias in aliases:
            if alias.upper() in upper_cols:
                col_map[upper_cols[alias.upper()]] = canonical
                break
    return df.rename(columns=col_map)


def _alias_score(row_vals: list) -> int:
    all_up = {a.upper() for aliases in _MASTER_ALIASES.values() for a in aliases}
    return sum(1 for v in row_vals if str(v).upper().strip() in all_up)


def _safe_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.upper() not in ("NAN", "NONE", "") else None


def _parse_master_xlsx(file_bytes: bytes) -> pd.DataFrame:
    """Return a normalized DataFrame from master XLSX with required columns."""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str, keep_default_na=False)
    scan = min(15, len(df_raw))
    header_row = max(range(scan), key=lambda i: _alias_score(df_raw.iloc[i].tolist()))
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row, dtype=str, keep_default_na=False)
    return _normalize_master_cols(df)


def _infer_producer(mnemonic: Optional[str]) -> str:
    """Komatsu if mnemonic starts with KOM, otherwise SCANIA (covers Scania + Hensley)."""
    if mnemonic and mnemonic.upper().startswith("KOM"):
        return "KOMATSU"
    return "SCANIA"


# ── GET /master/parts/meta ───────────────────────────────────────────────────

@router.get("/parts/meta")
async def get_master_meta(
    db: AsyncSession = Depends(get_db),
    _: Principal = Depends(require_user_role("admin")),
):
    result = await db.execute(
        select(MasterUpload, User)
        .join(User, User.id == MasterUpload.uploaded_by, isouter=True)
        .order_by(MasterUpload.uploaded_at.desc())
        .limit(1)
    )
    row = result.one_or_none()

    if row is None:
        # No master uploaded yet — return null-safe empty shape
        total_result = await db.execute(select(func.count(Part.id)).where(Part.is_active == True))
        total = total_result.scalar_one() or 0
        return {
            "filename": None,
            "uploaded_at": None,
            "uploader_name": None,
            "total": total,
            "class_v_count": 0,
            "class_g_count": 0,
            "komatsu_count": 0,
            "scania_count": 0,
        }

    upload, user = row
    return {
        "filename": upload.filename,
        "uploaded_at": upload.uploaded_at.isoformat() if upload.uploaded_at else None,
        "uploader_name": user.name if user else None,
        "total": upload.total_count,
        "class_v_count": upload.class_v_count,
        "class_g_count": upload.class_g_count,
        "komatsu_count": upload.komatsu_count,
        "scania_count": upload.scania_count,
    }


# ── GET /master/parts ────────────────────────────────────────────────────────

@router.get("/parts")
async def list_master_parts(
    limit: int = 50,
    page: int = 1,
    search: Optional[str] = None,
    kelas: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: Principal = Depends(require_user_role("admin")),
):
    filters = [Part.is_active == True]
    if kelas and kelas.upper() in ("V", "G"):
        filters.append(Part.kelas == kelas.upper())
    if search and search.strip():
        like = f"%{search.strip()}%"
        filters.append(
            or_(
                Part.part_number.ilike(like),
                Part.description.ilike(like),
                Part.mnemonic.ilike(like),
                Part.commodity.ilike(like),
            )
        )

    count_result = await db.execute(select(func.count(Part.id)).where(*filters))
    total = count_result.scalar_one() or 0

    result = await db.execute(
        select(Part)
        .where(*filters)
        .order_by(Part.part_number)
        .offset((page - 1) * limit)
        .limit(limit)
    )
    parts = result.scalars().all()

    return {
        "items": [
            {
                "part_number": p.part_number,
                "description": p.description,
                "mnemonic": p.mnemonic,
                "commodity": p.commodity,
                "kelas": p.kelas,
            }
            for p in parts
        ],
        "total": total,
    }


# ── POST /master/parts/upload ────────────────────────────────────────────────

@router.post("/parts/upload")
async def upload_master(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_user_role("admin")),
):
    filename = file.filename or "master.xlsx"
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only XLSX/XLS files are accepted")

    file_bytes = await file.read()
    if len(file_bytes) == 0:
        raise HTTPException(status_code=400, detail="File is empty")

    try:
        df = _parse_master_xlsx(file_bytes)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {e}")

    missing = _REQUIRED_MASTER - set(df.columns)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(sorted(missing))}. Found: {', '.join(df.columns.tolist())}",
        )

    inserted = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    class_v = 0
    class_g = 0
    komatsu = 0
    scania = 0
    now = datetime.now(timezone.utc)

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # type: ignore[arg-type]

        part_number = _safe_str(row.get("part_number"))
        if not part_number or part_number.upper() in ("PART NUMBER", "PART_NUMBER", "N/A", "-"):
            skipped += 1
            continue

        raw_kelas = _safe_str(row.get("kelas", ""))
        kelas = (raw_kelas or "V").upper()
        if kelas not in ("V", "G"):
            errors.append(f"Row {row_num}: Part {part_number} — invalid class '{raw_kelas}', expected V or G")
            skipped += 1
            continue

        description = _safe_str(row.get("description"))
        mnemonic = _safe_str(row.get("mnemonic"))
        commodity = _safe_str(row.get("commodity"))
        stockcode = _safe_str(row.get("stockcode"))
        producer = _infer_producer(mnemonic)

        try:
            existing = await db.execute(select(Part).where(Part.part_number == part_number))
            part = existing.scalar_one_or_none()

            if part is None:
                part = Part(
                    part_number=part_number,
                    description=description,
                    kelas=kelas,
                    mnemonic=mnemonic,
                    commodity=commodity,
                    stockcode=stockcode,
                    producer=producer,
                    is_active=True,
                )
                db.add(part)
                inserted += 1
            else:
                part.kelas = kelas
                part.is_active = True
                part.updated_at = now
                if description is not None:
                    part.description = description
                if mnemonic is not None:
                    part.mnemonic = mnemonic
                if commodity is not None:
                    part.commodity = commodity
                if stockcode is not None:
                    part.stockcode = stockcode
                part.producer = producer
                updated += 1

            await db.flush()

            if kelas == "V":
                class_v += 1
            else:
                class_g += 1

            if producer == "KOMATSU":
                komatsu += 1
            else:
                scania += 1

        except Exception as e:
            errors.append(f"Row {row_num}: Part {part_number} — {e}")

    total = class_v + class_g

    upload_record = MasterUpload(
        filename=filename,
        uploaded_by=principal.id,
        uploaded_at=now,
        total_count=total,
        class_v_count=class_v,
        class_g_count=class_g,
        komatsu_count=komatsu,
        scania_count=scania,
    )
    db.add(upload_record)
    await db.commit()

    return {
        "inserted": inserted,
        "updated": updated,
        "class_v": class_v,
        "class_g": class_g,
        "skipped": skipped,
        "errors": errors,
    }


# ── GET /master/parts/template ───────────────────────────────────────────────

@router.get("/parts/template")
async def download_master_template(
    _: Principal = Depends(require_user_role("admin")),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Class VG"

    headers = ["No", "Stockcode", "Part Number", "Description", "Mnemonic", "Class"]
    header_fill = PatternFill("solid", fgColor="1F6F4C")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Three sample rows
    samples = [
        [1, "KOM-ENG-001", "600-311-3750", "Filter Oli Engine Komatsu", "KOM-ENG", "V"],
        [2, "SCA-BDY-001", "1873018",      "Air Filter Scania P460",    "SCA-BDY", "G"],
        [3, "KOM-UDR-001", "207-70-73181", "Seal Kit Undercarriage",     "KOM-UDR", "V"],
    ]
    for sample in samples:
        ws.append(sample)

    # Column widths
    widths = [6, 18, 20, 45, 16, 8]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=master_class_vg_template.xlsx"},
    )
