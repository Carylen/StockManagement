import asyncio
import io
from datetime import date
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.database import get_db
from app.core.auth import require_role, Principal
from app.models.stock import StockLevel
from app.models.inquiry import Inquiry

router = APIRouter(prefix="/export", tags=["export"])

HEADER_FILL = PatternFill("solid", fgColor="F5A623")
HEADER_FONT = Font(bold=True, color="000000")


async def _make_xlsx_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    await asyncio.to_thread(wb.save, buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/inquiries")
async def export_inquiries(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role("user", "admin", "supplier")),
):
    result = await db.execute(
        select(Inquiry)
        .order_by(desc(Inquiry.created_at))
        .limit(1000)
    )
    inquiries = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inquiry Kelas G"

    headers = ["Tanggal", "Mekanik", "NRP", "Site",
               "Part Number", "Part Name", "Qty", "Status", "UT Notes", "Replacement PN"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for inq in inquiries:
        for item in inq.items:
            ws.cell(row=row_idx, column=1, value=inq.created_at.strftime("%d/%m/%Y") if inq.created_at else "")
            ws.cell(row=row_idx, column=2, value=inq.submitted_by_name or "")
            ws.cell(row=row_idx, column=3, value=inq.submitted_by_nrp or "")
            ws.cell(row=row_idx, column=4, value=inq.site)
            ws.cell(row=row_idx, column=5, value=item.part_number)
            ws.cell(row=row_idx, column=6, value=item.part_name or "")
            ws.cell(row=row_idx, column=7, value=item.qty)
            ws.cell(row=row_idx, column=8, value=item.status.upper())
            ws.cell(row=row_idx, column=9, value=item.ut_note or "")
            ws.cell(row=row_idx, column=10, value=item.replacement_pn or "")
            row_idx += 1

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    today = date.today().strftime("%Y%m%d")
    return await _make_xlsx_response(wb, f"inquiry_export_{today}.xlsx")


@router.get("/stock-report")
async def export_stock_report(
    db: AsyncSession = Depends(get_db),
    current_user: Principal = Depends(require_role("admin")),
):
    result = await db.execute(
        select(StockLevel)
        .where(StockLevel.site == current_user.site)
        .order_by(StockLevel.part_number)
    )
    stocks = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Stok {current_user.site}"

    headers = ["Part Number", "Deskripsi", "Komoditi", "RTT", "TBD", "Total", "MIN", "MAX", "Status", "Estimasi"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, stock in enumerate(stocks, 2):
        ws.cell(row=row_idx, column=1, value=stock.part_number)
        ws.cell(row=row_idx, column=2, value=stock.description or "")
        ws.cell(row=row_idx, column=3, value=stock.commodity or "")
        ws.cell(row=row_idx, column=4, value=stock.rtt_qty or 0)
        ws.cell(row=row_idx, column=5, value=stock.tbd_qty or 0)
        ws.cell(row=row_idx, column=6, value=(stock.rtt_qty or 0) + (stock.tbd_qty or 0))
        ws.cell(row=row_idx, column=7, value=float(stock.min_qty) if stock.min_qty is not None else 0)
        ws.cell(row=row_idx, column=8, value=float(stock.max_qty) if stock.max_qty is not None else 0)
        ws.cell(row=row_idx, column=9, value=stock.status or "")
        ws.cell(row=row_idx, column=10,
            value=stock.estimated_date.strftime("%d/%m/%Y") if stock.estimated_date else "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    today = date.today().strftime("%Y%m%d")
    return await _make_xlsx_response(wb, f"stok_{current_user.site}_{today}.xlsx")
