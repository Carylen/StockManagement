"""
Download template XLSX endpoints.

GET /templates/readiness  → template upload readiness harian (admin only)
GET /templates/master     → template master Class V/G (admin only)
GET /templates/employees  → template bulk karyawan (admin only)
"""
import asyncio
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from app.core.auth import require_user_role, Principal


router = APIRouter(prefix="/templates", tags=["templates"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _styled_header(ws, headers: list[str], bg_hex: str):
    """Write bold+colored header row, return the worksheet."""

    fill = PatternFill("solid", fgColor=bg_hex)
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _set_col_widths(ws, widths: list[int]):
    import openpyxl.utils
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w


# ── readiness template ────────────────────────────────────────────────────────

def _build_readiness(site: str | None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Readiness"

    headers = [
        "part_number", "description", "min", "max",
        "status", "rtt", "tbd", "total", "estimasi",
    ]
    _styled_header(ws, headers, "1F6F4C")

    today = str(date.today())
    samples = [
        ["600-311-3750", "Filter Oli Engine Komatsu",  2.0, 5.0, "AMAN",    4,  0, 4, 0],
        ["1873018",      "Air Filter Scania P460",      1.0, 3.0, "WARNING", 0,  1, 1, 2],
        ["207-70-73181", "Seal Kit Undercarriage",      1.0, 2.0, "MAX",     2,  0, 2, 0],
    ]
    for row in samples:
        ws.append(row)

    _set_col_widths(ws, [20, 40, 6, 6, 10, 6, 6, 8, 10])

    # Freeze header row
    ws.freeze_panes = "A2"

    # Info sheet
    info = wb.create_sheet("Info")
    notes = [
        ["Kolom", "Keterangan"],
        ["part_number", "Wajib. Nomor part (PN). Case-insensitive."],
        ["description", "Opsional. Nama/deskripsi part."],
        ["min", "Qty minimum stok (angka desimal diperbolehkan)."],
        ["max", "Qty maksimum stok. Harus >= min."],
        ["status", "WARNING | AMAN | OVER | MAX — di-recompute ulang oleh backend."],
        ["rtt", "Qty stok RTT (integer)."],
        ["tbd", "Qty stok TBD (integer)."],
        ["total", "Harus sama dengan rtt + tbd. Isi 0 jika tidak tahu."],
        ["estimasi", "Qty in-transit / estimasi kedatangan (integer)."],
        ["", ""],
        ["Catatan", f"Site: {site or '(sesuai akun admin yang upload)'}"],
    ]
    for r in notes:
        info.append(r)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── master template ───────────────────────────────────────────────────────────

def _build_master() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Class VG"

    headers = ["No", "Stockcode", "Part Number", "Description", "Mnemonic", "Class"]
    _styled_header(ws, headers, "1F6F4C")

    samples = [
        [1, "KOM-ENG-001", "600-311-3750",  "Filter Oli Engine Komatsu",  "KOM-ENG", "V"],
        [2, "SCA-BDY-001", "1873018",       "Air Filter Scania P460",     "SCA-BDY", "G"],
        [3, "KOM-UDR-001", "207-70-73181",  "Seal Kit Undercarriage",      "KOM-UDR", "V"],
    ]
    for row in samples:
        ws.append(row)

    _set_col_widths(ws, [6, 18, 20, 45, 16, 8])
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Info")
    notes = [
        ["Kolom",       "Keterangan"],
        ["No",          "Opsional. Nomor urut."],
        ["Stockcode",   "Opsional. Kode stok internal."],
        ["Part Number", "Wajib. Nomor part (PN)."],
        ["Description", "Opsional. Nama/deskripsi part."],
        ["Mnemonic",    "Opsional. Prefix KOM* → Komatsu, lainnya → Scania."],
        ["Class",       "Wajib. V atau G (huruf kapital)."],
    ]
    for r in notes:
        info.append(r)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── employees template ────────────────────────────────────────────────────────

def _build_employees() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Karyawan"

    headers = ["No", "NRP", "Nama", "Posisi", "Shift"]
    _styled_header(ws, headers, "1F6F4C")

    samples = [
        [1, "KM19142", "Budi Santoso",   "Mekanik", "Day"],
        [2, "KM20015", "Ahmad Fauzi",    "GL",       "Day"],
        [3, "KM18033", "Rini Widiastuti","Mekanik",  "Night"],
    ]
    for row in samples:
        ws.append(row)

    _set_col_widths(ws, [6, 14, 30, 12, 10])
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Info")
    notes = [
        ["Kolom",   "Keterangan"],
        ["No",      "Opsional. Nomor urut."],
        ["NRP",     "Wajib. ID karyawan (case-insensitive, disimpan uppercase)."],
        ["Nama",    "Wajib. Nama lengkap karyawan."],
        ["Posisi",  "Wajib. Mekanik atau GL."],
        ["Shift",   "Opsional. Day, Night, dsb."],
    ]
    for r in notes:
        info.append(r)
    info.column_dimensions["A"].width = 10
    info.column_dimensions["B"].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── routes ────────────────────────────────────────────────────────────────────

@router.get("/readiness")
async def download_readiness_template(
    principal: Principal = Depends(require_user_role("admin")),
):
    try:
        data = await asyncio.to_thread(_build_readiness, principal.site)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))

    site_tag = f"_{principal.site}" if principal.site else ""
    filename = f"template_readiness{site_tag}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/master")
async def download_master_template(
    _: Principal = Depends(require_user_role("admin")),
):
    try:
        data = await asyncio.to_thread(_build_master)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))

    return StreamingResponse(
        io.BytesIO(data),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=template_master_class_vg.xlsx"},
    )


@router.get("/employees")
async def download_employees_template(
    _: Principal = Depends(require_user_role("admin")),
):
    try:
        data = await asyncio.to_thread(_build_employees)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))

    return StreamingResponse(
        io.BytesIO(data),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": "attachment; filename=template_karyawan.xlsx"},
    )
