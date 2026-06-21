import asyncio
import io
import math
import os
import uuid
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, func, case, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.auth import Principal
from app.utils.permissions import require_permission, require_any_permission
from app.models.plan_period import PlanPeriod
from app.models.plan_line import PlanLine
from app.models.plan_line_history import PlanLineHistory
from app.models.plan_revision import PlanRevision
from app.models.plan_scope_seen import PlanScopeSeen
from app.models.plan_upload_session import PlanUploadSession
from app.models.permission import SupplierSite
from app.schemas.plan import (
    MergeResult, EventCreateResult, PeriodListItem, PlanLineCreateRequest, PlanLineOut,
    PaginatedLines, FillRequest, OverviewResponse, AchievementResponse, HistoryItem,
    FillImportResult, CoordinationItem, RevisionRequest, RevisionResponse, SeenRequest,
    UploadDiffSummary, UploadPreviewRow, UploadSessionResult, AttentionResponse,
)
from app.services.plan_service import (
    parse_and_merge_plan_file, parse_and_diff_plan_file, merge_plan_lines, period_state,
    aggregate_period, record_history, parse_fill_file, list_accessible_periods,
    _row_to_dict, _row_from_dict, now_wib,
)
from app.services.plan_parser import ACTIVITIES
from app.services.plan_collaboration_service import (
    to_line_out, build_coordination, derive_readiness, origin_visibility_clause,
)
from app.services.plan_attention_service import build_attention
from app.services.excel_templates import build_plan_template, XLSX_MIME
from app.utils.scoping import has_all_sites

router = APIRouter(prefix="/scheduled-plans", tags=["scheduled-plans"])

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
FILL_IMPORT_EXTENSIONS = {".xlsx", ".xls", ".csv"}
UPLOAD_SESSION_TTL = timedelta(minutes=30)

_HEADER_FILL = PatternFill("solid", fgColor="E8A323")
_HEADER_FONT = Font(bold=True, color="000000")


def _no_active_event() -> HTTPException:
    """Structured 404 for 'event belum ada' — single source so every endpoint
    that loads a period by id reports the same {code, message} shape instead
    of a plain string (DELTA3 D.1)."""
    return HTTPException(status_code=404, detail={
        "code": "NO_ACTIVE_EVENT",
        "message": "Belum ada event aktif dari admin untuk periode ini.",
    })


async def _supplier_period_or_403(period_id: str, principal: Principal, db: AsyncSession) -> PlanPeriod:
    """Load a period and ensure the supplier is assigned to its site."""
    period = (await db.execute(select(PlanPeriod).where(PlanPeriod.id == period_id))).scalar_one_or_none()
    if period is None:
        raise _no_active_event()
    assigned = (await db.execute(
        select(SupplierSite.id).where(
            SupplierSite.supplier_id == principal.id,
            SupplierSite.site_code == period.site,
        )
    )).scalar_one_or_none()
    if assigned is None:
        raise HTTPException(status_code=403, detail="Site tidak ter-assign ke supplier ini")
    return period


def _check_extension(filename: str) -> bool:
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_EXTENSIONS


async def _get_period_scoped(period_id: str, principal: Principal, db: AsyncSession) -> PlanPeriod:
    """Load a period and enforce site scoping (own-site unless can_view_all_sites)."""
    period = (await db.execute(select(PlanPeriod).where(PlanPeriod.id == period_id))).scalar_one_or_none()
    if period is None:
        raise _no_active_event()
    if not has_all_sites(principal) and period.site != principal.site:
        raise HTTPException(status_code=403, detail="Event di luar scope site Anda")
    return period


async def _to_period_item(db: AsyncSession, period: PlanPeriod, show_pct: bool) -> PeriodListItem:
    """Single-period readiness snapshot (BASELINE lines only)."""
    total, ready = (await db.execute(
        select(
            func.count(),
            func.sum(case((PlanLine.is_ready.is_(True), 1), else_=0)),
        ).where(
            PlanLine.period_id == period.id,
            PlanLine.removed_in_revision.is_(False),
            PlanLine.origin == PlanLine.ORIGIN_BASELINE,
        )
    )).one()
    total, ready = int(total or 0), int(ready or 0)
    pct = round(ready / total * 100, 1) if total else 0.0
    return PeriodListItem(
        period_id=period.id, site=period.site, name=period.name,
        start_date=period.start_date, due_date=period.due_date,
        state=period_state(period.due_date),
        readiness_pct=pct if show_pct else None, total_lines=total,
    )


def _merge_result(m) -> MergeResult:
    return MergeResult(
        rows_total=m.rows_total, rows_inserted=m.rows_inserted,
        rows_updated=m.rows_updated, rows_merged=m.rows_merged, errors=m.errors,
    )


def _days_remaining(due_date: date) -> int:
    return (due_date - now_wib().date()).days


# ── 3.0 Create event + baseline upload (Admin) ────────────────────────────
@router.post("/periods", response_model=EventCreateResult)
async def create_event(
    name: str = Form(...),
    start_date: date = Form(...),
    due_date: date = Form(...),
    site: str | None = Form(None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_plan_event")),
):
    """Admin creates a scheduled-plan event with explicit dates and uploads the
    agreed baseline in one step. Planner cannot upload into an activity until
    this exists."""
    target_site = site if has_all_sites(principal) else principal.site
    if not target_site:
        raise HTTPException(status_code=422, detail="Site wajib diisi")
    if due_date <= start_date:
        raise HTTPException(status_code=422, detail="due_date harus setelah start_date")

    dup = (await db.execute(
        select(PlanPeriod.id).where(PlanPeriod.site == target_site, PlanPeriod.name == name)
    )).scalar_one_or_none()
    if dup is not None:
        raise HTTPException(status_code=409, detail=f"Event '{name}' sudah ada untuk site {target_site}")

    if not _check_extension(file.filename or ""):
        raise HTTPException(status_code=400, detail="Hanya file XLSX yang diterima")
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File kosong")

    period = PlanPeriod(
        id=str(uuid.uuid4()), site=target_site, name=name,
        start_date=start_date, due_date=due_date, uploaded_by=principal.id,
    )
    db.add(period)
    await db.flush()

    merge = await parse_and_merge_plan_file(
        file_bytes=file_bytes, filename=file.filename or "plan.xlsx",
        period=period, origin=PlanLine.ORIGIN_BASELINE, actor_id=principal.id, db=db,
    )
    await db.commit()

    return EventCreateResult(period=await _to_period_item(db, period, show_pct=True), merge=_merge_result(merge))


# ── 3.0b Add more to the baseline (Admin) ─────────────────────────────────
@router.post("/periods/{period_id}/baseline-upload", response_model=MergeResult)
async def baseline_upload(
    period_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_plan_event")),
):
    """Admin adds more agreed items to an existing event's baseline."""
    period = await _get_period_scoped(period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")
    if not _check_extension(file.filename or ""):
        raise HTTPException(status_code=400, detail="Hanya file XLSX yang diterima")
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File kosong")

    merge = await parse_and_merge_plan_file(
        file_bytes=file_bytes, filename=file.filename or "plan.xlsx",
        period=period, origin=PlanLine.ORIGIN_BASELINE, actor_id=principal.id, db=db,
    )
    await db.commit()
    return _merge_result(merge)


# ── 3.1 Upload extra items into an existing event (Planner) ──────────────
@router.post("/periods/{period_id}/upload", response_model=MergeResult)
async def upload_plan(
    period_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_scheduled_plan")),
):
    """Planner uploads into an event admin already created. Rows matching an
    existing (baseline or extra) line just update it; brand-new rows are
    inserted as EXTRA — visible only to admin and this planner."""
    period = await _get_period_scoped(period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")
    if not _check_extension(file.filename or ""):
        raise HTTPException(status_code=400, detail="Hanya file XLSX yang diterima")
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File kosong")

    merge = await parse_and_merge_plan_file(
        file_bytes=file_bytes, filename=file.filename or "plan.xlsx",
        period=period, origin=PlanLine.ORIGIN_EXTRA, actor_id=principal.id, db=db,
    )
    await db.commit()
    return _merge_result(merge)


# ── 3.1a Preview/diff upload before commit (Planner) — two-step, lazy-expire
@router.post("/periods/{period_id}/upload/preview", response_model=UploadSessionResult)
async def upload_plan_preview(
    period_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_scheduled_plan")),
):
    """Parse + diff the file against the event's current lines WITHOUT
    writing to plan_lines. Stores the result as a PENDING session the planner
    must /confirm (or /discard) within 30 minutes — confirm re-runs the exact
    same merge_plan_lines() the direct /upload endpoint uses."""
    period = await _get_period_scoped(period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")
    if not _check_extension(file.filename or ""):
        raise HTTPException(status_code=400, detail="Hanya file XLSX yang diterima")
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File kosong")

    diff, parse = await parse_and_diff_plan_file(
        file_bytes=file_bytes, filename=file.filename or "plan.xlsx", period=period, db=db,
    )

    # Lazy cleanup: drop this planner's own stale PENDING sessions for this
    # event before creating a new one — no cron, just expire-on-next-use.
    now = datetime.now(timezone.utc)
    await db.execute(delete(PlanUploadSession).where(
        PlanUploadSession.period_id == period_id,
        PlanUploadSession.uploaded_by == principal.id,
        PlanUploadSession.status == PlanUploadSession.STATUS_PENDING,
        PlanUploadSession.expires_at < now,
    ))

    all_rows = diff.inserted + [r for _, r in diff.updated]
    session = PlanUploadSession(
        id=str(uuid.uuid4()), period_id=period_id, uploaded_by=principal.id,
        source_filename=file.filename or "plan.xlsx",
        diff_payload={
            "origin": PlanLine.ORIGIN_EXTRA,
            "rows": [_row_to_dict(r) for r in all_rows],
            "errors": parse.errors,
        },
        status=PlanUploadSession.STATUS_PENDING,
        expires_at=now + UPLOAD_SESSION_TTL,
    )
    db.add(session)
    await db.commit()

    rows_preview = [
        UploadPreviewRow(
            action="INSERT", egi=r.egi, cn=r.cn, apl_activity=r.apl_activity, npn=r.npn,
            description=r.description or None, req_qty=r.req_qty, req_date=r.req_date,
        ) for r in diff.inserted[:20]
    ] + [
        UploadPreviewRow(
            action="UPDATE", egi=r.egi, cn=r.cn, apl_activity=r.apl_activity, npn=r.npn,
            description=r.description or None, req_qty=r.req_qty, req_date=r.req_date,
        ) for _, r in diff.updated[:20]
    ]

    return UploadSessionResult(
        session_id=session.id,
        summary=UploadDiffSummary(
            inserted=len(diff.inserted), updated=len(diff.updated),
            marked_removed=0, errors=parse.errors,
        ),
        rows_preview=rows_preview[:20],
    )


async def _load_own_session(session_id: str, principal: Principal, db: AsyncSession) -> PlanUploadSession:
    session = (await db.execute(
        select(PlanUploadSession).where(PlanUploadSession.id == session_id)
    )).scalar_one_or_none()
    if session is None or session.uploaded_by != principal.id:
        raise HTTPException(status_code=404, detail="Sesi upload tidak ditemukan")
    return session


@router.post("/upload-sessions/{session_id}/confirm", response_model=MergeResult)
async def confirm_upload_session(
    session_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_scheduled_plan")),
):
    session = await _load_own_session(session_id, principal, db)
    if session.status != PlanUploadSession.STATUS_PENDING:
        raise HTTPException(status_code=409, detail="Sesi upload sudah diproses")
    now = datetime.now(timezone.utc)
    if session.expires_at < now:
        session.status = PlanUploadSession.STATUS_DISCARDED
        await db.commit()
        raise HTTPException(status_code=410, detail="Sesi upload sudah kedaluwarsa, silakan upload ulang")

    period = await _get_period_scoped(session.period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")

    rows = [_row_from_dict(d) for d in session.diff_payload["rows"]]
    merge = await merge_plan_lines(
        period=period, rows=rows,
        origin=session.diff_payload.get("origin", PlanLine.ORIGIN_EXTRA),
        actor_id=principal.id, db=db,
    )
    merge.errors = list(session.diff_payload.get("errors", []))
    period.source_filename = session.source_filename
    period.revised_at = now
    session.status = PlanUploadSession.STATUS_CONFIRMED
    await db.commit()
    return _merge_result(merge)


@router.post("/upload-sessions/{session_id}/discard", status_code=204)
async def discard_upload_session(
    session_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_scheduled_plan")),
):
    session = await _load_own_session(session_id, principal, db)
    if session.status == PlanUploadSession.STATUS_PENDING:
        session.status = PlanUploadSession.STATUS_DISCARDED
        await db.commit()


# ── 3.1b Add a single line manually (Admin → BASELINE, Planner → EXTRA) ──
@router.post("/periods/{period_id}/lines", response_model=PlanLineOut)
async def add_line(
    period_id: str,
    body: PlanLineCreateRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission("can_manage_plan_event", "can_manage_scheduled_plan")),
):
    period = await _get_period_scoped(period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")

    activity = body.activity.strip().upper()
    if activity not in ACTIVITIES:
        raise HTTPException(status_code=422, detail=f"ACTIVITY harus salah satu dari: {', '.join(sorted(ACTIVITIES))}")
    origin = PlanLine.ORIGIN_BASELINE if "can_manage_plan_event" in principal.permissions else PlanLine.ORIGIN_EXTRA

    egi, cn, npn, apl = body.egi.strip().upper(), body.cn.strip().upper(), body.npn.strip().upper(), body.apl_activity.strip().upper()
    dup = (await db.execute(select(PlanLine.id).where(
        PlanLine.period_id == period_id, PlanLine.egi == egi, PlanLine.cn == cn,
        PlanLine.npn == npn, PlanLine.apl_activity == apl,
    ))).scalar_one_or_none()
    if dup is not None:
        raise HTTPException(status_code=409, detail="Baris dengan EGI/CN/NPN/APL ACTIVITY ini sudah ada di event ini")

    line = PlanLine(
        id=str(uuid.uuid4()), period_id=period_id, activity=activity,
        egi=egi, cn=cn, npn=npn, apl_activity=apl,
        description=(body.description or None), req_qty=Decimal(str(body.req_qty)),
        req_date=body.req_date, origin=origin, created_by=principal.id, updated_by=principal.id,
    )
    db.add(line)
    await db.commit()
    await db.refresh(line)
    return to_line_out(line, mask_origin="can_view_plan_achievement" not in principal.permissions)


# ── 3.2 List periods/events (Planner / Admin / Supplier) ─────────────────
@router.get("/periods", response_model=list[PeriodListItem])
async def list_periods(
    site: str | None = None,
    page: int = 1,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_view_plan_achievement", "can_manage_plan_event", "can_fill_scheduled_plan")),
):
    accessible = await list_accessible_periods(db, principal, site=site)
    periods = accessible[(page - 1) * limit: (page - 1) * limit + limit]
    perms = principal.permissions

    if not periods:
        return []

    # live readiness aggregation, BASELINE only (ignore removed/EXTRA lines)
    agg = await db.execute(
        select(
            PlanLine.period_id,
            func.count().label("total"),
            func.sum(case((PlanLine.is_ready.is_(True), 1), else_=0)).label("ready"),
        )
        .where(
            PlanLine.period_id.in_([p.id for p in periods]),
            PlanLine.removed_in_revision.is_(False),
            PlanLine.origin == PlanLine.ORIGIN_BASELINE,
        )
        .group_by(PlanLine.period_id)
    )
    stats = {row.period_id: (int(row.total), int(row.ready or 0)) for row in agg}

    # Achievement % is admin-only — planner/supplier never receive it, even
    # though they can otherwise see the periods themselves.
    show_pct = "can_view_plan_achievement" in perms

    out = []
    for p in periods:
        total, ready = stats.get(p.id, (0, 0))
        pct = round(ready / total * 100, 1) if total else 0.0
        out.append(PeriodListItem(
            period_id=p.id, site=p.site, name=p.name,
            start_date=p.start_date, due_date=p.due_date,
            state=period_state(p.due_date),
            readiness_pct=pct if show_pct else None, total_lines=total,
        ))
    return out


# ── 3.3 List lines of an event (Planner / Admin read-only) ───────────────
@router.get("/periods/{period_id}/lines", response_model=PaginatedLines)
async def list_lines(
    period_id: str,
    activity: str | None = None,
    apl_activity: str | None = None,
    status: str | None = Query(None, pattern="^(READY|NOT_READY)$"),
    include_removed: bool = False,
    include_extra: bool = True,
    page: int = 1,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_view_plan_achievement")),
):
    period = (await db.execute(select(PlanPeriod).where(PlanPeriod.id == period_id))).scalar_one_or_none()
    if period is None:
        raise HTTPException(status_code=404, detail="Event tidak ditemukan")

    base = select(PlanLine).where(
        PlanLine.period_id == period_id,
        origin_visibility_clause(principal, include_extra=include_extra),
    )
    if not include_removed:
        base = base.where(PlanLine.removed_in_revision.is_(False))
    if activity:
        base = base.where(PlanLine.activity == activity.upper())
    if apl_activity:
        base = base.where(PlanLine.apl_activity == apl_activity)
    if status:
        base = base.where(PlanLine.status == status)

    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one() or 0
    rows = (await db.execute(
        base.order_by(PlanLine.apl_activity, PlanLine.egi, PlanLine.cn, PlanLine.npn)
        .offset((page - 1) * limit).limit(limit)
    )).scalars().all()

    # Planner shouldn't learn which of their own lines got flagged EXTRA —
    # that signal exists for admin to raise with them, not to dodge.
    is_admin = "can_view_plan_achievement" in principal.permissions
    return PaginatedLines(
        items=[to_line_out(r, mask_origin=not is_admin) for r in rows],
        total=total, page=page, limit=limit,
        pages=math.ceil(total / limit) if total else 1,
    )


# ── 3.5 Fill list (UT/Supplier) — BASELINE only ───────────────────────────
@router.get("/fill", response_model=PaginatedLines)
async def list_fill_lines(
    period_id: str,
    apl_activity: str | None = None,
    status: str | None = Query(None, pattern="^(READY|NOT_READY)$"),
    page: int = 1,
    limit: int = 200,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_fill_scheduled_plan")),
):
    period = (await db.execute(select(PlanPeriod).where(PlanPeriod.id == period_id))).scalar_one_or_none()
    if period is None:
        raise HTTPException(status_code=404, detail="Event tidak ditemukan")

    assigned = (await db.execute(
        select(SupplierSite.id).where(
            SupplierSite.supplier_id == principal.id,
            SupplierSite.site_code == period.site,
        )
    )).scalar_one_or_none()
    if assigned is None:
        raise HTTPException(status_code=403, detail="Site tidak ter-assign ke supplier ini")

    base = select(PlanLine).where(
        PlanLine.period_id == period_id,
        PlanLine.removed_in_revision.is_(False),
        PlanLine.origin == PlanLine.ORIGIN_BASELINE,
    )
    if apl_activity:
        base = base.where(PlanLine.apl_activity == apl_activity)
    if status:
        base = base.where(PlanLine.status == status)

    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one() or 0
    rows = (await db.execute(
        base.order_by(PlanLine.apl_activity, PlanLine.egi, PlanLine.cn, PlanLine.npn)
        .offset((page - 1) * limit).limit(limit)
    )).scalars().all()

    return PaginatedLines(
        items=[to_line_out(r) for r in rows],
        total=total, page=page, limit=limit,
        pages=math.ceil(total / limit) if total else 1,
    )


# ── 3.5 Fill by UT/Supplier — readiness derived from ut_location + est_date
@router.patch("/lines/{line_id}/fill", response_model=PlanLineOut)
async def fill_line(
    line_id: str,
    body: FillRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_fill_scheduled_plan")),
):
    line = (await db.execute(select(PlanLine).where(PlanLine.id == line_id))).scalar_one_or_none()
    if line is None:
        raise HTTPException(status_code=404, detail="Baris tidak ditemukan")
    if line.origin != PlanLine.ORIGIN_BASELINE:
        raise HTTPException(status_code=403, detail="Item ini belum termasuk baseline yang disetujui")

    period = (await db.execute(select(PlanPeriod).where(PlanPeriod.id == line.period_id))).scalar_one_or_none()
    if period is None:
        raise HTTPException(status_code=404, detail="Event tidak ditemukan")

    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")

    # supplier must be assigned to the period's site
    assigned = (await db.execute(
        select(SupplierSite.id).where(
            SupplierSite.supplier_id == principal.id,
            SupplierSite.site_code == period.site,
        )
    )).scalar_one_or_none()
    if assigned is None:
        raise HTTPException(status_code=403, detail="Site tidak ter-assign ke supplier ini")

    status, is_ready = derive_readiness(body.ut_location, body.est_date)

    # audit field-level changes before mutating
    record_history(db, line.id, "status", line.status, status, principal.id)
    record_history(db, line.id, "ut_location", line.ut_location, body.ut_location, principal.id)
    record_history(db, line.id, "est_date", line.est_date, body.est_date, principal.id)

    line.status = status
    line.ut_location = body.ut_location
    line.est_date = body.est_date
    line.is_ready = is_ready
    line.updated_by = principal.id

    await db.commit()
    await db.refresh(line)
    return to_line_out(line)


# ── 3.6a Export fill template (UT/Supplier) — no Status column ───────────
@router.get("/fill/export")
async def export_fill(
    period_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_fill_scheduled_plan")),
):
    period = await _supplier_period_or_403(period_id, principal, db)

    rows = (await db.execute(
        select(PlanLine)
        .where(
            PlanLine.period_id == period_id,
            PlanLine.removed_in_revision.is_(False),
            PlanLine.origin == PlanLine.ORIGIN_BASELINE,
        )
        .order_by(PlanLine.apl_activity, PlanLine.egi, PlanLine.cn, PlanLine.npn)
    )).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fill"
    # Rows are matched on re-upload by natural key (EGI, CN, APL ACTIVITY, NPN)
    # — the same key used everywhere else in this module — not an internal id.
    headers = ["EGI", "CN", "APL ACTIVITY", "NPN", "DESC", "REQ QTY", "REQ DATE", "UT LOCATION", "EST DATE"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, ln in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=ln.egi)
        ws.cell(row=i, column=2, value=ln.cn)
        ws.cell(row=i, column=3, value=ln.apl_activity)
        ws.cell(row=i, column=4, value=ln.npn)
        ws.cell(row=i, column=5, value=ln.description or "")
        ws.cell(row=i, column=6, value=float(ln.req_qty) if ln.req_qty is not None else 0)
        ws.cell(row=i, column=7, value=ln.req_date.strftime("%d/%m/%Y") if ln.req_date else "")
        ws.cell(row=i, column=8, value=ln.ut_location or "")
        ws.cell(row=i, column=9, value=ln.est_date.strftime("%d/%m/%Y") if ln.est_date else "")

    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 45)
    ws.freeze_panes = "A2"

    # Info sheet — site/due-date context (DELTA3 D.2: this response is a file
    # stream, not JSON, so the days-remaining passthrough lives here).
    info = wb.create_sheet("Info")
    info.append(["Site", period.site])
    info.append(["Event", period.name])
    info.append(["Due date", period.due_date.strftime("%d/%m/%Y")])
    info.append(["Sisa hari sebelum LOCKED", _days_remaining(period.due_date)])
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    await asyncio.to_thread(wb.save, buf)
    buf.seek(0)
    fname = f"fill_{period.name}_{period.site}_{date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── 3.6b Import filled file (UT/Supplier) — readiness derived from ut_location + est_date
@router.post("/fill/import", response_model=FillImportResult)
async def import_fill(
    period_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_fill_scheduled_plan")),
):
    _, ext = os.path.splitext((file.filename or "").lower())
    if ext not in FILL_IMPORT_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Hanya file XLSX atau CSV yang diterima")

    period = await _supplier_period_or_403(period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="File kosong")

    parsed = await asyncio.to_thread(parse_fill_file, file_bytes, file.filename or "fill.xlsx")
    if parsed["error"]:
        raise HTTPException(status_code=422, detail=parsed["error"])

    # index this period's live BASELINE lines by natural key — EXTRA lines
    # are never fillable, and the supplier never adds/removes rows this way.
    lines = (await db.execute(
        select(PlanLine).where(
            PlanLine.period_id == period_id,
            PlanLine.removed_in_revision.is_(False),
            PlanLine.origin == PlanLine.ORIGIN_BASELINE,
        )
    )).scalars().all()
    by_key = {(ln.egi, ln.cn, ln.apl_activity, ln.npn): ln for ln in lines}

    updated = 0
    skipped = 0
    errors: list[dict] = list(parsed["errors"])
    for r in parsed["rows"]:
        line = by_key.get((r.egi, r.cn, r.apl_activity, r.npn))
        if line is None:
            skipped += 1
            errors.append({
                "egi": r.egi, "cn": r.cn, "apl_activity": r.apl_activity, "npn": r.npn,
                "reason": "EGI/CN/APL ACTIVITY/NPN tidak ditemukan di event ini",
            })
            continue
        status, is_ready = derive_readiness(r.ut_location, r.est_date)
        changed = False
        changed |= record_history(db, line.id, "status", line.status, status, principal.id)
        changed |= record_history(db, line.id, "ut_location", line.ut_location, r.ut_location, principal.id)
        changed |= record_history(db, line.id, "est_date", line.est_date, r.est_date, principal.id)
        line.status = status
        line.ut_location = r.ut_location
        line.est_date = r.est_date
        line.is_ready = is_ready
        line.updated_by = principal.id
        if changed:
            updated += 1
        else:
            skipped += 1

    await db.commit()
    return FillImportResult(
        updated=updated, skipped=skipped, errors=errors,
        end_date=period.due_date, days_remaining=_days_remaining(period.due_date),
    )


# ── 3.4b Batch revision per apl_activity (Planner) ───────────────────────
@router.post("/periods/{period_id}/revisions", response_model=RevisionResponse)
async def create_revision(
    period_id: str,
    body: RevisionRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_manage_scheduled_plan")),
):
    period = await _get_period_scoped(period_id, principal, db)
    if period_state(period.due_date) == "LOCKED":
        raise HTTPException(status_code=403, detail="Event sudah LOCKED, tidak bisa diubah")

    # revision_no = max+1 per (period, apl_activity)
    last_no = (await db.execute(
        select(func.max(PlanRevision.revision_no)).where(
            PlanRevision.period_id == period_id,
            PlanRevision.apl_activity == body.apl_activity,
        )
    )).scalar_one_or_none()
    revision_no = (last_no or 0) + 1

    updated = 0
    for item in body.lines:
        line = (await db.execute(select(PlanLine).where(PlanLine.id == item.line_id))).scalar_one_or_none()
        # only touch lines that belong to this period & apl_activity scope
        if line is None or line.period_id != period_id or line.apl_activity != body.apl_activity:
            continue
        record_history(db, line.id, "req_date", line.req_date, item.req_date, principal.id)
        line.req_date = item.req_date
        line.updated_by = principal.id
        updated += 1

    db.add(PlanRevision(
        period_id=period_id,
        apl_activity=body.apl_activity,
        revision_no=revision_no,
        note=body.note,
        revised_by=principal.id,
    ))
    await db.commit()
    return RevisionResponse(revision_no=revision_no, updated_lines=updated)


# ── 3.4c Coordination summary per apl_activity (Planner or Supplier) ─────
@router.get("/periods/{period_id}/coordination", response_model=list[CoordinationItem])
async def coordination(
    period_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_fill_scheduled_plan")),
):
    is_planner = "can_manage_scheduled_plan" in principal.permissions
    if is_planner:
        period = await _get_period_scoped(period_id, principal, db)
    else:
        period = await _supplier_period_or_403(period_id, principal, db)
    return await build_coordination(
        db, period, viewer_id=principal.id,
        viewer_side="planner" if is_planner else "supplier",
    )


# ── 3.4d Mark a scope as seen (clears unread badge) ──────────────────────
@router.post("/periods/{period_id}/seen", status_code=204)
async def mark_seen(
    period_id: str,
    body: SeenRequest,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_fill_scheduled_plan")),
):
    # ensure access (planner=site scope, supplier=assignment)
    if "can_manage_scheduled_plan" in principal.permissions:
        await _get_period_scoped(period_id, principal, db)
    else:
        await _supplier_period_or_403(period_id, principal, db)

    row = (await db.execute(
        select(PlanScopeSeen).where(
            PlanScopeSeen.user_id == principal.id,
            PlanScopeSeen.period_id == period_id,
            PlanScopeSeen.apl_activity == body.apl_activity,
        )
    )).scalar_one_or_none()
    now = datetime.now(timezone.utc)
    if row is None:
        db.add(PlanScopeSeen(
            user_id=principal.id, period_id=period_id,
            apl_activity=body.apl_activity, last_seen_at=now,
        ))
    else:
        row.last_seen_at = now
    await db.commit()


# ── Line history (audit trail) ───────────────────────────────────────────
@router.get("/lines/{line_id}/history", response_model=list[HistoryItem])
async def line_history(
    line_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_view_plan_achievement")),
):
    line = (await db.execute(select(PlanLine).where(PlanLine.id == line_id))).scalar_one_or_none()
    if line is None:
        raise HTTPException(status_code=404, detail="Baris tidak ditemukan")
    await _get_period_scoped(line.period_id, principal, db)  # site scope guard

    rows = (await db.execute(
        select(PlanLineHistory)
        .where(PlanLineHistory.line_id == line_id)
        .order_by(PlanLineHistory.changed_at.desc())
    )).scalars().all()
    return [HistoryItem.model_validate(r) for r in rows]


# ── Attention digest (cross-role) — DELTA3 section C ──────────────────────
@router.get("/attention", response_model=AttentionResponse)
async def attention_digest(
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_fill_scheduled_plan",
        "can_view_plan_achievement", "can_manage_plan_event")),
):
    items = await build_attention(db, principal)
    return AttentionResponse(items=items)


# ── Empty template download (planner/admin baseline vs supplier fill) ────
@router.get("/template")
async def plan_template(
    role: str = Query(..., pattern="^(planner|supplier)$"),
    principal: Principal = Depends(require_any_permission(
        "can_manage_scheduled_plan", "can_manage_plan_event", "can_fill_scheduled_plan")),
):
    data = await asyncio.to_thread(build_plan_template, role)
    fname = f"scheduled_plan_template_{role}.xlsx"
    return StreamingResponse(
        io.BytesIO(data), media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── 3.7 Overview (Admin) — achievement % is admin-only, BASELINE only ────
@router.get("/overview", response_model=OverviewResponse)
async def plan_overview(
    period_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_view_plan_achievement")),
):
    period = await _get_period_scoped(period_id, principal, db)
    activities = await aggregate_period(db, period)
    return OverviewResponse(period_id=period.id, activities=activities)


# ── 3.8 Achievement (Admin Site) — BASELINE only ──────────────────────────
@router.get("/achievement", response_model=AchievementResponse)
async def plan_achievement(
    period_id: str,
    db: AsyncSession = Depends(get_db),
    principal: Principal = Depends(require_permission("can_view_plan_achievement")),
):
    period = await _get_period_scoped(period_id, principal, db)
    activities = await aggregate_period(db, period)
    return AchievementResponse(
        period_id=period.id,
        activities=[{
            "activity": act["activity"],
            "readiness_pct": act["readiness_pct"],
            "ready": act["ready"],
            "total": act["total"],
            "not_ready_apl_activities": [a for a in act["apl_activities"] if a["pct"] < 100],
        } for act in activities],
    )
